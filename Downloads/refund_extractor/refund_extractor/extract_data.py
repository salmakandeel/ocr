"""
Usage:
    python3 extract_data.py "/path/to/استردادات" [output_draft.xlsx]

Walks the given root folder. Each immediate subfolder is treated as one
refund case and is expected to contain exactly two images:
  - a bank account-confirmation letter (any format/bank)
  - the university refund-request form (with student name + amount
    handwritten in)

Images are told apart by CONTENT (OCR keyword scoring), not filename, so
they can be named anything.

Produces one workbook with two sheets:
  - "Data"   — exactly the columns you specified, ready for the bank file.
               Transaction Amount and Remittance Information are filled
               with the OCR's best guess but are UNRELIABLE (handwritten
               fields) — always check them against the "Review" sheet.
  - "Review" — one row per case with the cropped handwriting snippets
               embedded as images, the OCR guess, and the source folder,
               so you can fix the Data sheet quickly without reopening the
               original photos.

After you've checked/corrected the "Data" sheet, run build_batches.py on
it to split into files of 25 rows named after their Transaction Amount sum.
"""
import sys
import re
from pathlib import Path

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from ocr_utils import classify_image
from extract_bank_letter import extract_bank_letter_from_text
from extract_university_form import extract_university_form_from_image

IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".bmp", ".tif", ".tiff", ".webp"}

# ---- Fixed / constant values you gave me -----------------------------
DEBTOR_NAME = "جامعة المنوفيه الاهليه"
DEBTOR_ACCOUNT_NUMBER = "20314180200"
DEBTOR_ACCOUNT_TYPE = "CACC"
TRANSACTION_PURPOSE = "cash"

COLUMNS = [
    "Instruction ID", "Creditor Name", "Creditor Account Number",
    "Creditor Account Type", "Creditor Bank", "Creditor Bank Branch",
    "Debtor Name", "Debtor Account Number", "Debtor Account Type",
    "Transaction Amount", "Transaction Purpose", "Remittance Information",
    "Creditor National ID",
]

REVIEW_ROW_HEIGHT = 60
REVIEW_IMG_WIDTH_PX = 260


def find_case_folders(root: Path):
    return sorted([p for p in root.iterdir() if p.is_dir()])


def find_images(folder: Path):
    return sorted([p for p in folder.iterdir() if p.suffix.lower() in IMAGE_EXTS])


def clean_amount(raw: str):
    """Best-effort turn an OCR guess like 'املع 22, © 5' or '50,225' into a
    plain number string. Returns '' if nothing digit-like was found —
    that's a strong signal the row needs manual entry."""
    digits = re.sub(r"[^\d]", "", raw or "")
    return digits


def process_case(folder: Path):
    images = find_images(folder)
    row = {c: "" for c in COLUMNS}
    review = {
        "folder": folder.name,
        "name_crop": None, "name_guess": "",
        "amount_crop": None, "amount_guess": "",
        "warnings": [],
    }

    if len(images) != 2:
        review["warnings"].append(
            f"Expected 2 images, found {len(images)} — skipped, check this folder manually."
        )
        return row, review

    kinds = {}
    ocr_cache = {}  # path -> (kind, text, oriented_image)
    for img_path in images:
        kind, text, oriented_im = classify_image(str(img_path))
        ocr_cache[img_path] = (kind, text, oriented_im)
        kinds.setdefault(kind, []).append(img_path)

    bank_path = kinds.get("bank_letter", [None])[0]
    form_path = kinds.get("university_form", [None])[0]

    # Fallback if classification was ambiguous for one of them but not both
    if bank_path is None or form_path is None:
        remaining = [p for p in images if p not in (bank_path, form_path)]
        if bank_path is None and remaining:
            bank_path = remaining.pop(0)
        if form_path is None and remaining:
            form_path = remaining.pop(0)
        review["warnings"].append("Could not confidently tell the two images apart by content — verify assignment.")

    if bank_path:
        _kind, bank_text, _im = ocr_cache[bank_path]
        bank_data = extract_bank_letter_from_text(bank_text)
        row["Creditor Name"] = bank_data["creditor_name"]
        row["Creditor Account Number"] = bank_data["creditor_account_no"]
        row["Creditor Account Type"] = bank_data["creditor_account_type"]
        row["Creditor Bank"] = bank_data["creditor_bank"]
        if not bank_data["creditor_name"]:
            review["warnings"].append("Creditor name not found — check bank letter.")
        if not bank_data["creditor_account_no"]:
            review["warnings"].append("Account number/IBAN not found — check bank letter.")
        if not bank_data["creditor_bank"]:
            review["warnings"].append("Bank abbreviation not recognized — add it to bank_map.json.")

    if form_path:
        _kind, _text, form_im = ocr_cache[form_path]
        form_data = extract_university_form_from_image(form_im)
        review["name_crop"] = form_data["student_name_crop"]
        review["name_guess"] = form_data["student_name_guess"]
        review["amount_crop"] = form_data["amount_crop"]
        review["amount_guess"] = form_data["amount_guess"]

        row["Remittance Information"] = form_data["student_name_guess"]
        row["Transaction Amount"] = clean_amount(form_data["amount_guess"])
        review["warnings"].append(
            "Student name & amount are HANDWRITTEN — OCR guess is unreliable, verify against the snippet images."
        )

    row["Debtor Name"] = DEBTOR_NAME
    row["Debtor Account Number"] = DEBTOR_ACCOUNT_NUMBER
    row["Debtor Account Type"] = DEBTOR_ACCOUNT_TYPE
    row["Transaction Purpose"] = TRANSACTION_PURPOSE
    row["Creditor National ID"] = ""  # left blank, as requested
    row["Creditor Bank Branch"] = ""

    return row, review


def build_workbook(rows, reviews, out_path: Path):
    wb = openpyxl.Workbook()

    # ---- Data sheet ----
    ws = wb.active
    ws.title = "Data"
    ws.append(COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for i, row in enumerate(rows, start=1):
        row["Instruction ID"] = i
        ws.append([row[c] for c in COLUMNS])
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(14, len(col_name) + 2)

    # Flag amount cells that came back empty (couldn't OCR any digits at all)
    warn_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    amount_col = COLUMNS.index("Transaction Amount") + 1
    for i in range(2, ws.max_row + 1):
        cell = ws.cell(row=i, column=amount_col)
        if not str(cell.value).strip():
            cell.fill = warn_fill

    # ---- Review sheet ----
    rv = wb.create_sheet("Review")
    rv.append(["Row #", "Folder", "Student name (snippet)", "OCR guess (name)",
               "Amount (snippet)", "OCR guess (amount)", "Warnings"])
    for cell in rv[1]:
        cell.font = Font(bold=True)
    rv.column_dimensions["B"].width = 20
    rv.column_dimensions["C"].width = 38
    rv.column_dimensions["D"].width = 22
    rv.column_dimensions["E"].width = 26
    rv.column_dimensions["F"].width = 18
    rv.column_dimensions["G"].width = 50

    tmp_dir = out_path.parent / "_tmp_review_imgs"
    tmp_dir.mkdir(exist_ok=True)

    for i, rev in enumerate(reviews, start=1):
        r = i + 1
        rv.cell(row=r, column=1, value=i)
        rv.cell(row=r, column=2, value=rev["folder"])
        rv.cell(row=r, column=4, value=rev["name_guess"])
        rv.cell(row=r, column=6, value=rev["amount_guess"])
        rv.cell(row=r, column=7, value=" | ".join(rev["warnings"]))
        rv.row_dimensions[r].height = REVIEW_ROW_HEIGHT

        if rev["name_crop"] is not None:
            p = tmp_dir / f"name_{i}.png"
            rev["name_crop"].save(p)
            img = XLImage(str(p))
            ratio = REVIEW_IMG_WIDTH_PX / img.width
            img.width = REVIEW_IMG_WIDTH_PX
            img.height = int(img.height * ratio)
            rv.add_image(img, f"C{r}")

        if rev["amount_crop"] is not None:
            p = tmp_dir / f"amount_{i}.png"
            rev["amount_crop"].save(p)
            img = XLImage(str(p))
            ratio = 160 / img.width
            img.width = 160
            img.height = int(img.height * ratio)
            rv.add_image(img, f"E{r}")

    wb.save(out_path)
    print(f"Wrote draft workbook: {out_path}  ({len(rows)} case(s))")
    print(f"NOTE: temporary snippet images live in {tmp_dir} — the workbook has its own copies embedded, "
          f"so you can delete that folder once you're happy with the file, but don't delete it before then.")


def main():
    if len(sys.argv) < 2:
        print("Usage: python3 extract_data.py \"/path/to/استردادات\" [output_draft.xlsx]")
        sys.exit(1)
    root = Path(sys.argv[1])
    out_path = Path(sys.argv[2]) if len(sys.argv) > 2 else root.parent / "draft_output.xlsx"

    case_folders = find_case_folders(root)
    if not case_folders:
        print(f"No subfolders found under {root}")
        sys.exit(1)

    rows, reviews = [], []
    for folder in case_folders:
        print(f"Processing: {folder.name} ...")
        row, review = process_case(folder)
        rows.append(row)
        reviews.append(review)

    build_workbook(rows, reviews, out_path)


if __name__ == "__main__":
    main()
